# import pandas as pd
import argparse
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import os
import shutil
import sys
from file_logger import FileLogger


class AssetFormFiller:
    """固定资产领用/回收单填写工具"""

    def __init__(self, template_path, output_dir=None, log=None):
        """
        初始化
        """
        self.template_path = template_path
        self.output_dir =output_dir if output_dir else "forms"
        self.log = log

        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

        if not self.log:
            self.log = FileLogger('logs', 'assetFormFiller')

    def parse_data_from_file(self, file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data_text = f.read()
                data = self.parse_tab_data(data_text)
                return data
        except Exception as e:
            self.log.error(f"读取文件失败: {e}")
            sys.exit(1)

    def parse_tab_data(self, text_data):
        """
        解析制表符分隔的资产数据

        输入格式：
        使用人	部门	分类	操作类型	资产名称	资产编码	品牌型号	备注
        周明军	文创城法庭	其他	回收	台式计算机	510122MB1867896218000450	戴尔OptiPlex 3050 SFF 003068	文创城法庭	根据统一安排替换，从文创城法庭回收
        """
        lines = text_data.strip().split('\n')
        assets = []

        for line in lines:
            if not line.strip():
                continue

            # 按制表符分割
            parts = line.split('\t')
            if len(parts) >= 7:
                asset = {
                    'user': parts[0].strip(),  # 使用人
                    'department': parts[1].strip(),  # 部门
                    'reason': parts[2].strip(),  # 事由
                    'action': parts[3].strip(),  # 操作类型（回收/发放）
                    'name': parts[4].strip(),  # 资产名称
                    'asset_code': parts[5].strip(),  # 资产编码
                    'brand_model': parts[6].strip(),  # 品牌型号
                    'location':parts[7].strip() if len(parts) > 7 else '',  #地点
                    'remark': parts[8].strip() if len(parts) > 8 else '',  # 备注
                    'asset_type': parts[9].strip() if len(parts) > 9 else '固定资产'  #资产类型
                }
                assets.append(asset)

        return assets

    def group_assets_by_user(self, assets):
        """按使用人分组资产数据"""
        grouped = {}
        for asset in assets:
            if '搬迁' in asset['action']:
                continue
            user_key = f"{asset['asset_type']}_{asset['user']}_{asset['department']}"
            if user_key not in grouped:
                grouped[user_key] = {
                    'asset_type':asset['asset_type'],
                    'user': asset['user'],
                    'department': asset['department'],
                    'assets': []
                }
            grouped[user_key]['assets'].append(asset)

        return grouped


    def fill_form(self, form_data, output_filename=None):
        """
        填写领用/回收单

        form_data 格式:
        {
            'form_type': '领用',  # 或 '回收' 或 ’搬迁’
            'department': '部门',
            'user': '姓名',
            'date': '2026-05-12',
            'reason': '工作需要',
            'assets': [
                {
                    'user': '使用人',
                    'department': '部门',
                    'reason': '事由',
                    'action': '操作类型（回收/发放）',
                    'name': '资产名称',
                    'asset_code': '资产编码',
                    'brand_model': '品牌型号',
                    'location':'地点',
                    'remark': '备注'
                    'asset_type': '资产类型'
                },
                # ...
            ]
        }
        """
        try:

            # 确定输出文件名
            if output_filename is None:
                wr_date = datetime.now().strftime('%Y%m%d')
                output_filename = f"{wr_date}{form_data.get('asset_type', '资产')}申领单-{form_data.get('user', 'unknown')}.xlsx"

            output_path = os.path.join(self.output_dir, output_filename)

            # 加载模板
            if self.template_path and os.path.exists(self.template_path):
                shutil.copy(self.template_path, output_path)
                wb = load_workbook(output_path)
                ws = wb.active
            else:
                # 没有模板则创建新文件
                wb = Workbook()
                ws = wb.active
                ws.title = f"{form_data.get('asset_type', '固定资产')}领用回收单"

            # 1. 设置标题（如果是新创建的模板）
            if not self.template_path:
                ws.merge_cells('A1:H1')
                ws['A1'] = f"{form_data.get('asset_type', '固定资产')}{form_data.get('form_type', '领用/回收')}单"
                ws['A1'].font = Font(name='宋体', size=16, bold=True)
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            else:
                # 修改已有模板的标题
                title_cell = ws['A1']
                title_cell.value = f"{form_data.get('asset_type', '固定资产')}{form_data.get('form_type', '领用/回收')}单"

            # 2. 填写申请部门（第2行，C列）
            department_cell = ws['C2']
            department_cell.value = form_data.get('department', '')
            department_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

            # 3. 填写使用人（第2行，G列）
            # user_cell = ws['G2']
            user_cell = ws['F2']
            user_cell.value = form_data.get('user', '')
            user_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

            # 4. 填写领用/回收时间（第3行，C列）
            # date_cell = ws['C3']
            date_cell = ws['H2']
            date_cell.value = form_data.get('date', datetime.now().strftime('%Y-%m-%d'))

            # 5. 填写资产数据
            assets = form_data.get('assets', [])
            # start_row = 6  # 数据起始行
            start_row = 5

            for idx, asset in enumerate(assets):
                row = start_row + idx
                if row > 10:  # 最多5行
                    self.log.warning(f"⚠ 资产数量超过5行，第{idx + 1}行及以后将不被显示")
                    break

                # # 序号
                # seq_cell = ws.cell(row=row, column=1, value=idx + 1)
                # seq_cell.alignment = Alignment(horizontal='center', vertical='center')
                # seq_cell.border = thin_border

                # 资产名称
                name_cell = ws.cell(row=row, column=2, value=asset.get('name', ''))
                name_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                # 品牌型号
                model_cell = ws.cell(row=row, column=4, value=asset.get('brand_model', ''))
                model_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                # 安装地点
                location_cell = ws.cell(row=row, column=6, value=asset.get('location', ''))
                location_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                # 资产编码
                code_cell = ws.cell(row=row, column=7, value=asset.get('asset_code', ''))
                code_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                # 备注（领用/回收标识）
                remark_text = f"{asset.get('action', '')} {asset.get('remark', '')}"
                remark_cell = ws.cell(row=row, column=8, value=remark_text)
                remark_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

            # 6. 填写申请事由
            reason_cell = ws['C10']
            reason_cell.value = form_data.get('reason', '')

            # 保存文件
            wb.save(output_path)
            # print(f"✓ 表单已生成: {output_path}")
            return output_path

        except Exception as e:
            self.log.error(f"✗ 填写表单失败: {e}")
            import traceback
            traceback.print_exc()
            return None


    def fill_forms(self, data_list):
        # 按使用人分组
        grouped = self.group_assets_by_user(data_list)

        self.log.debug(f"✅ 分组为 {len(grouped)} 个表单")

        # 为每个使用人生成表单
        results = []
        for user_key, group in grouped.items():
            self.log.debug(f"为{group['asset_type']} {group['user']} ({group['department']}) 生成表单")
            output_file = self.fill_form(group)
            if output_file:
                results.append(output_file)
                self.log.info(f"✅ 表单已保存: {output_file}")
        self.log.info(f"✅ 总共生成 {len(results)} 个表单")



def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='资产领用/回收单生成工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
    使用示例:
  # 直接传入制表符分隔的数据
  python asset_form.py "姓名\t部门\t事由\t发放/回收\t资产名称\t资产编码\t资产型号\t地点\t备注"
  
  # 传入多行数据（用引号包裹）
  python asset_form.py "姓名\t部门\t事由\t发放/回收\t资产名称\t资产编码1\t资产型号\t地点\t备注
姓名\t部门\t事由\t发放/回收\t资产名称\t资产编码2\t资产型号\t地点\t备注"
  
  # 从文件读取数据
  python asset_form.py --file assets.txt
  
  # 指定输出文件名
  python asset_form.py --data "姓名\t部门\t事由\t发放/回收\t资产名称\t资产编码\t资产型号\t地点\t备注" --output 周明军_资产表单.xlsx
        '''
    )

    parser.add_argument('data', nargs='?', help='资产数据（制表符分隔）')
    parser.add_argument('--file', '-f',
                        default="./in_file.txt",help='从文件读取数据')
    parser.add_argument('--template-path', '-t',
                        default="./固定资产领用及回收单-模板.xlsx", help='指定模板文件')
    parser.add_argument('--output', '-o', help='输出文件名')
    parser.add_argument('--output-dir', '-d', default='forms', help='输出目录（默认: forms）')
    parser.add_argument('--verbose', '-v', action='store_true', help='显示详细信息')

    args = parser.parse_args()

    # 获取数据
    data_text = None

    if args.file:
        # 从文件读取
        try:
            with open(args.file, 'r', encoding='utf-8') as f:
                data_text = f.read()
            if args.verbose:
                print(f"✓ 从文件读取: {args.file}")
        except Exception as e:
            print(f"读取文件失败: {e}")
            sys.exit(1)
    elif args.data:
        # 从命令行参数读取
        data_text = args.data
        if args.verbose:
            print(f"✓ 从命令行读取数据")
    else:
        # 没有提供数据，显示帮助
        parser.print_help()
        sys.exit(1)

    if not data_text or not data_text.strip():
        print("错误：没有提供数据")
        sys.exit(1)

    filler = AssetFormFiller(template_path=args.template_path, output_dir=args.output_dir)
    print("\n正在解析数据...")
    assets = filler.parse_tab_data(data_text)
    print(f"✓ 成功解析 {len(assets)} 条资产记录")

    if not assets:
        print("错误：无法解析数据，请确保格式正确（使用制表符分隔）")
        print("格式：使用人\t部门\t分类\t操作类型\t资产名称\t资产编码\t品牌型号\t备注")
        sys.exit(1)

    if args.verbose:
        print(f"✓ 解析到 {len(assets)} 条资产记录")
        for i, asset in enumerate(assets, 1):
            print(f"  {i}. {asset['name']} - {asset['action']} - {asset['user']}")


    # 按使用人分组
    grouped = filler.group_assets_by_user(assets)

    if args.verbose:
        print(f"\n✓ 分组为 {len(grouped)} 个表单")

    # 为每个使用人生成表单
    results = []
    for user_key, group in grouped.items():
        if args.verbose:
           print(f"\n为{group['asset_type']} {group['user']} ({group['department']}) 生成表单...")
        output_file = filler.fill_form(group)
        if output_file:
            results.append(output_file)
            # print(f"  ✓ 表单已保存: {output_file}")

    if not args.verbose:
        print(f"\n✓ 成功生成 {len(results)} 个表单")
        for f in results:
            print(f"  - {f}")


if __name__ == "__main__":
    main()