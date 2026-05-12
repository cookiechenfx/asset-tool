# 这是一个示例 Python 脚本。

# 按 Shift+F10 执行或将其替换为您的代码。
# 按 双击 Shift 在所有地方搜索类、文件、工具窗口、操作和设置。
#
#
# def print_hi(name):
#     # 在下面的代码行中使用断点来调试脚本。
#     print(f'Hi, {name}')  # 按 Ctrl+F8 切换断点。
#
#
# # 按装订区域中的绿色按钮以运行脚本。
# if __name__ == '__main__':
#     from selenium import webdriver
#     from selenium.webdriver.chrome.options import Options
#     from selenium.webdriver.chrome.service import Service
#
#     # 设置无头模式
#     chrome_options = Options()
#     chrome_options.add_argument('--headless')  # 关键：不显示界面
#     chrome_options.add_argument('--no-sandbox')
#     chrome_options.add_argument('--disable-dev-shm-usage')  # 解决资源有限导致的崩溃
#
#     # 指定 Chromium 路径（通常无需指定，若报错则取消注释）
#     chrome_options.binary_location = '/usr/bin/chromium-browser'
#
#     service = Service('/usr/lib/chromium-browser/chromedriver')
#
#     driver = webdriver.Chrome(service=service, options=chrome_options)
#     driver.set_page_load_timeout(30)
#
#     print("浏览器驱动初始化成功")
#
#     # 你的业务代码...
#     driver.get("https://gz.dothantech.com/#/login")
#     print(driver.title)
#
#     driver.quit()
import time
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side


class AssetManagementSystem:
    def __init__(self, username, password, login_url):
        """初始化资产管理系统脚本"""
        self.username = username
        self.password = password
        self.login_url = login_url
        self.driver = None
        self.wait = None

        # 设置Chrome选项
        chrome_options = Options()
        chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        # 如果不使用debug模式，注释上一行，启用下一行
        # chrome_options.add_argument("--start-maximized")

        self.init_driver(chrome_options)

    def init_driver(self, chrome_options):
        """初始化浏览器驱动"""
        """初始化浏览器驱动 - Ubuntu 版本"""
        try:
            chrome_options = Options()

            # 无头模式配置
            # chrome_options.add_argument('--headless')  # 无界面运行
            chrome_options.add_argument('--no-sandbox')  # 解决权限问题
            chrome_options.add_argument('--disable-dev-shm-usage')  # 解决资源限制
            chrome_options.add_argument('--disable-gpu')  # 禁用GPU加速
            chrome_options.add_argument('--window-size=1920,1080')  # 设置窗口大小

            # 指定 Chromium 路径（Ubuntu 默认路径）
            chrome_options.binary_location = '/usr/bin/chromium-browser'

            # 指定 Chromedriver 路径
            service = Service('/usr/lib/chromium-browser/chromedriver')

            # 创建驱动
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.driver.set_page_load_timeout(30)
            self.wait = WebDriverWait(self.driver, 10)

            print("浏览器驱动初始化成功")
            return True
        except Exception as e:
            print(f"初始化失败: {e}")
            print("\n请确保已安装: sudo apt install chromium-browser chromium-chromedriver")
            return False

    def login(self):
        """登录系统"""
        try:
            self.driver.get(self.login_url)
            time.sleep(2)

            # 获取用户名和密码输入框（请根据实际情况修改选择器）
            # username_input = self.wait.until(
            #     EC.presence_of_element_located((By.ID, "username"))
            # )
            password_input = self.driver.find_element(By.XPATH, "//input[@type='password']")
            username_input = self.driver.find_element(By.XPATH, "//input[@type='text']")
            # password_input = self.driver.find_element(By.NAME, "password")

            # password_input = self.driver.find_element(By.ID, "password")

            # 输入登录信息
            username_input.send_keys(self.username)
            password_input.send_keys(self.password)

            # 点击登录按钮
            login_btn = self.driver.find_element(By.XPATH, "//button[@type='button']")
            login_btn.click()

            time.sleep(3)
            print("登录成功")
            return True
        except Exception as e:
            print(f"登录失败: {e}")
            return False

    def navigate_to_asset_list(self):
        """导航到资产列表页面"""
        try:
            print("\n正在导航到资产列表...")

            # 等待侧边栏加载
            time.sleep(2)

            # 方法1：先展开"资产管理"父菜单，再点击"资产列表"
            parent_title = self.wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "//div[@class='el-submenu__title']//span[contains(text(), '资产管理')]/.."))
            )

            # 通过父菜单的兄弟元素查找
            parent_li = parent_title.find_element(By.XPATH, "./..")  # 找到li.el-submenu

            # 检查父菜单是否已展开（子菜单是否可见）
            sub_menu = parent_li.find_element(By.XPATH, "//ul[@role='menu']")

            # 如果子菜单是隐藏的，点击展开
            if sub_menu and 'display: none' in sub_menu.get_attribute('style'):
                # 点击父菜单展开
                self.driver.execute_script("arguments[0].click()", parent_li)
                print("✓ 已展开'资产管理'菜单")
                time.sleep(1)

            # 点击"资产列表"
            asset_list = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), '资产列表')]"))
            )
            self.driver.execute_script("arguments[0].scrollIntoView(true);", asset_list)
            time.sleep(0.5)
            asset_list.click()
            print("✓ 通过父菜单+子菜单方式点击资产列表")
            time.sleep(2)
            return True

        except Exception as e:
            print(f"✗ 导航失败: {e}")
            return False

    def search_asset(self, asset_code):
        """搜索资产 - 适配通用搜索框"""
        try:
            print(f"\n搜索资产: {asset_code}")
            time.sleep(2)

            selector = "//input[@class='el-input__inner' and @placeholder='资产名称/编码/规格型号/供应商']"

            search_input = None
            try:
                search_input = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, selector))
                )
                print(f"✓ 找到搜索框: {selector}")
            except Exception as e:
                print(f"未找到搜索框: {e}")


            if not search_input:
                print("✗ 未找到搜索框")
                return False

            # 清空并输入资产编码
            search_input.clear()
            time.sleep(0.5)
            search_input.send_keys(asset_code)
            print(f"✓ 已输入搜索内容: {asset_code}")

            # 等待搜索建议或直接搜索
            time.sleep(1)

            # 查找查询按钮
            selector = "//div[contains(@class, 'el-input-group__append')]//button//i[@class='el-icon-search']/parent::button"

            try:
                query_btn = self.driver.find_element(By.XPATH, selector)
                if query_btn.is_displayed() and query_btn.is_enabled():
                    query_btn.click()
                    print("✓ 已点击查询按钮")
                    btn_found = True
            except Exception as e:
                    print(f"未找到搜索按钮: {e}")


            if not btn_found:
                # 按回车搜索
                search_input.send_keys("\n")
                print("✓ 已按回车搜索")

            # 等待搜索结果加载
            time.sleep(3)

            # 等待表格或结果出现
            try:
                self.wait.until(EC.presence_of_element_located((By.XPATH, "//table")))
                print("✓ 搜索结果已加载")
            except:
                print("⚠ 未检测到表格，但继续执行")

            return True

        except Exception as e:
            print(f"✗ 搜索资产失败: {e}")
            return False

    def select_asset(self, asset_code):
        """选中资产"""
        try:
            print(f"\n准备选中资产: {asset_code}")
            # time.sleep(2)

            # code_selector = "//input[@class='el-input__inner' and @placeholder='资产名称/编码/规格型号/供应商']"
            # code_selector = "//td[contains(@class='el-table_1_column4')]"
            # name_selector = "//td[@class='el-table_1_column5']"
            # code_selector = "//table//tbody//tr/td[4]"
            row_selector = f"//td[contains(@class, 'el-table_1_column_4')]//div[contains(text(), '{asset_code}')]/ancestor::tr"

            row = self.driver.find_element(By.XPATH, row_selector)

            print(f"✓ 找到资产所在行")

            checkbox = None
            checkbox_selector = f".//td[contains(@class, 'el-table_1_column_1')]//input[@type='checkbox']"
            # checkbox_selector = f".//td[contains(@class, 'el-table_1_column_1')]//label[contains(@class, 'el-checkbox')]"

            try:
                elem = row.find_element(By.XPATH, checkbox_selector)
                self.driver.execute_script("arguments[0].click();", elem)
                print("✓ 通过JavaScript点击input元素成功")
                time.sleep(0.5)
                return True
            except Exception as e:
                print(f"选中复选框失败：{e}")
                return False

        except Exception as e:
            print(f"✗ 选中资产失败: {e}")
            return False

    def modify_asset(self, asset_code, modified_mode, modified_data):
        """修改固定资产台账"""

        try:
            # 搜索需要修改的资产
            if not self.search_asset(asset_code):
                print("无法找到指定资产")
                return False

            if not self.select_asset(asset_code):
                print("无法选中指定资产")
                return False

            time.sleep(2)

            # 点击操作按钮
            action_btn = None
            action_selector = "//button[contains(.,'操作')]"
            action_btn = self.driver.find_element(By.XPATH, action_selector)
            action_btn.click()
            time.sleep(1)

            manage_selector = ""
            if "发放" in modified_mode:
                manage_selector = f"//li[contains(@class, 'el-dropdown-menu__item') and contains(., '领用')]"

            elif "回收" in modified_mode:
                manage_selector = f"//li[contains(@class, 'el-dropdown-menu__item') and contains(., '归还')]"

            elif "搬迁" in modified_mode:
                manage_selector = f"//li[contains(@class, 'el-dropdown-menu__item') and contains(., '变更')]"
            else:
                print(f"未知的调整模式：{modified_mode}")
                return False
            btn = self.driver.find_element(By.XPATH, manage_selector)
            btn.click()
            time.sleep(1)

            # 修改数据
            if "回收" in modified_mode:
                selector = f"//div[contains(@class, 'el-form-item')]//input[contains(@class, 'el-input__inner') and contains(@placeholder, '请输入或选择')]"
                # elem = self.driver.find_element(By.XPATH, selector)
                elem = None
                try:
                    elem = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, selector))
                    )
                    print(f"✓ 找到地址输入框: {selector}")
                except Exception as e:
                    print(f"未找到地址输入框: {e}")

                if not elem:
                    print("✗ 未找到地址输入框")
                    return False

                # 清空并输入资产编码
                elem.clear()
                time.sleep(0.5)
                elem.send_keys(modified_data['location'])
                print(f"✓ 已填写内容: {modified_data['location']}")
                time.sleep(2)

                # confirm_btn = self.wait.until(
                #     EC.element_to_be_clickable((By.XPATH,
                #                                 "//div[contains(@class, 'el-dialog__footer')]//button[contains(@class, 'el-button--primary')]//span[contains(text(), '确')]/parent::button"))
                # )
                # self.driver.execute_script("arguments[0].click();", confirm_btn)
                # print("✓ JS点击成功")

                selector = f"//div[contains(@class,'dialog-footer')]//button[contains(., '确')]"
                btn = self.driver.find_element(By.XPATH, selector)

                # try:
                #     elem = self.wait.until(
                #         EC.presence_of_element_located((By.XPATH, selector))
                #     )
                #     print(f"✓ 找到确定按钮: {selector}")
                # except Exception as e:
                #     print(f"未找到确定按钮: {e}")
                #
                # if not elem:
                #     print("✗ 未找到确定按钮")
                #     return False

                try:
                    self.driver.execute_script("arguments[0].click();", btn)
                    print("✓ JS点击成功")
                    time.sleep(2)
                except:
                    return False

                # btn.click()
            # elif "发放" in modified_mode:
            #     manage_selector = f"//li[contains(@class, 'el-dropdown-menu__item') and contains(., '归还')]"
            #
            # elif "搬迁" in modified_mode:
            #     manage_selector = f"//li[contains(@class, 'el-dropdown-menu__item') and contains(., '变更')]"
            else:
                print(f"未知的调整模式：{modified_mode}")
                return False

            # for field, value in modified_data.items():
            #     try:
            #         input_element = self.driver.find_element(By.ID, field)
            #         input_element.clear()
            #         input_element.send_keys(value)
            #     except:
            #         print(f"未找到字段: {field}")
            #
            # # 保存修改
            # save_btn = self.driver.find_element(By.XPATH, "//button[contains(text(), '保存')]")
            # save_btn.click()
            # time.sleep(2)

            print(f"资产 {asset_code} 修改成功")
            return True
        except Exception as e:
            print(f"修改资产失败: {e}")
            return False

    def batch_modify_assets(self, modification_list):
        """批量修改固定资产"""
        results = []
        for asset in modification_list:
            status = self.modify_asset(asset['asset_code'], asset['modified_data'])
            results.append({
                'asset_code': asset['asset_code'],
                'status': '成功' if status else '失败'
            })
            time.sleep(1)

        return results

    def generate_recovery_form(self, asset_info, form_type='recovery'):
        """生成领用/回收单

        参数:
        asset_info: 资产信息字典
        form_type: 表单类型 - 'recovery'(回收) 或 'receive'(领用)
        """
        wb = Workbook()
        ws = wb.active

        # 设置标题
        title = "固定资产领用单" if form_type == 'receive' else "固定资产回收单"
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = title
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置表头样式
        header_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 填写表单内容
        form_data = [
            ["单据编号", asset_info.get('form_no', self.generate_form_no()), "", "申请日期",
             asset_info.get('apply_date', datetime.now().strftime('%Y-%m-%d')), "", ""],
            ["申请部门", asset_info.get('department', ''), "", "申请人", asset_info.get('applicant', ''), "", ""],
            ["资产编码", asset_info.get('asset_code', ''), "", "资产名称", asset_info.get('asset_name', ''), "", ""],
            ["规格型号", asset_info.get('model', ''), "", "购置日期", asset_info.get('purchase_date', ''), "", ""],
            ["资产原值", asset_info.get('original_value', ''), "", "累计折旧", asset_info.get('depreciation', ''), "",
             ""],
            ["使用状态", asset_info.get('status', ''), "", "存放地点", asset_info.get('location', ''), "", ""],
            ["", "", "", "", "", "", ""],
            ["备注", asset_info.get('remarks', ''), "", "", "", "", ""],
            ["", "", "", "", "", "", ""],
            ["申请人签字", "", "日期", "", "部门负责人", "", ""],
            ["", "", "", "", "", "", ""],
            ["资产管理员", "", "日期", "", "", "", ""]
        ]

        # 填写数据
        for row_idx, row_data in enumerate(form_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15

        # 保存文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{title}_{asset_info.get('asset_code', 'unknown')}_{timestamp}.xlsx"
        wb.save(filename)
        print(f"表单已生成: {filename}")
        return filename

    def generate_form_no(self):
        """生成表单编号"""
        return f"GD{datetime.now().strftime('%Y%m%d')}{datetime.now().strftime('%H%M%S')}"

    def export_asset_list(self, search_condition=None):
        """导出资产列表"""
        try:
            if search_condition:
                # 输入搜索条件
                for key, value in search_condition.items():
                    try:
                        input_element = self.driver.find_element(By.ID, key)
                        input_element.clear()
                        input_element.send_keys(value)
                    except:
                        pass

                # 查询
                search_btn = self.driver.find_element(By.XPATH, "//button[contains(text(), '查询')]")
                search_btn.click()
                time.sleep(2)

            # 导出Excel
            export_btn = self.driver.find_element(By.XPATH, "//button[contains(text(), '导出')]")
            export_btn.click()
            time.sleep(3)

            print("资产列表导出成功")
            return True
        except Exception as e:
            print(f"导出失败: {e}")
            return False

    def close(self):
        """关闭浏览器"""
        if self.driver:
            self.driver.quit()
            print("浏览器已关闭")


def main():
    """主函数 - 使用示例"""

    # 配置信息
    config = {
        'username': '13683498313',  # 替换为您的用户名
        'password': 'tfxxhzcgl619%',  # 替换为您的密码
        'login_url': 'https://gz.dothantech.com/#/login'  # 替换为您的系统登录地址
    }

    # 初始化系统
    ams = AssetManagementSystem(config['username'], config['password'], config['login_url'])

    try:
        # 登录
        if not ams.login():
            print("登录失败，程序退出")
            return

        # 示例1：修改单个资产
        modified_data = {
            'asset_name': '台式计算机',
            'user': '张三',  # 使用人
            'location': '测试库房',  # 存放地点
            'code': '510116MB1867896226000061',
            'mode':'回收'
        }
        ams.navigate_to_asset_list()
        # ams.search_asset(modified_data['code'])
        ams.modify_asset(modified_data['code'], modified_data['mode'], modified_data)

        # # 示例2：批量修改资产
        # modifications = [
        #     {
        #         'asset_code': 'ASSET001',
        #         'modified_data': {'user': '张三', 'location': 'A栋301室'}
        #     },
        #     {
        #         'asset_code': 'ASSET002',
        #         'modified_data': {'user': '李四', 'location': 'B栋202室'}
        #     }
        # ]
        # results = ams.batch_modify_assets(modifications)
        # print("批量修改结果:", results)

        # 示例3：生成领用单
        asset_info = {
            'form_no': 'GD20231201001',
            'apply_date': '2023-12-01',
            'department': '信息技术部',
            'applicant': '王五',
            'asset_code': 'CP-2023-001',
            'asset_name': '笔记本电脑',
            'model': 'ThinkPad X1 Carbon',
            'purchase_date': '2023-06-01',
            'original_value': '12,000.00',
            'depreciation': '500.00',
            'status': '领用中',
            'location': 'A栋1001室',
            'remarks': '办公使用'
        }

        # 生成领用单
        receive_form = ams.generate_recovery_form(asset_info, 'receive')

        # 生成回收单
        asset_info['remarks'] = '设备更新，原设备回收'
        recovery_form = ams.generate_recovery_form(asset_info, 'recovery')

        # 示例4：导出现有资产列表
        # ams.export_asset_list({'department': '信息技术部'})

    except Exception as e:
        print(f"程序运行出错: {e}")

    finally:
        # 关闭浏览器
        ams.close()


# 使用配置文件的方式
def load_config():
    """从配置文件加载系统配置"""
    config_file = 'asset_config.json'
    import json
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except FileNotFoundError:
        print("配置文件不存在，创建默认配置...")
        default_config = {
            "username": "your_username",
            "password": "your_password",
            "login_url": "http://your-system-url/login",
            "chrome_driver_path": "chromedriver.exe"
        }
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(default_config, f, ensure_ascii=False, indent=2)
        print(f"请填写配置文件 {config_file} 后再运行")
        return None


if __name__ == "__main__":
    # 方式1：直接运行
    main()

    # 方式2：从配置文件加载
    # config = load_config()
    # if config:
    #     ams = AssetManagementSystem(
    #         config['username'],
    #         config['password'],
    #         config['login_url']
    #     )
    #     # ... 后续操作
