from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import os
import shutil
import sys
from asset_automator import AssetAutomator
from file_logger import FileLogger

class AssetExporter:
    """固定资产领用/回收单填写工具"""

    def __init__(self, log, output_dir=None):
        """
        初始化
        """
        # self.asset_list = asset_list
        self.output_dir = output_dir if output_dir else "result"
        self.log = log

        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)


    def load_data(self, in_file_path):
        try:
            with open(in_file_path, 'r', encoding='utf-8') as f:
                data_text = f.read()

                lines = data_text.strip().split('\n')
                staffs = []

                for line in lines:
                    if not line.strip():
                        continue

                    # 按制表符分割
                    parts = line.split('\t')
                    if len(parts) >= 2:
                        staff = {
                            'user': parts[0].strip(),  # 使用人
                            'department': parts[1].strip(),  # 部门
                        }
                        staffs.append(staff)
                return staffs
        except Exception as e:
            self.log.error(f"读取文件失败: {e}")
            sys.exit(1)


    def write_excel(self, data_list, output_filename=None):
        """
        输出查询到的资产结果

        """
        try:

            # 确定输出文件名
            if output_filename is None:
                wr_date = datetime.now().strftime('%Y-%m-%d %H.%M.%S')
                output_filename = f"{wr_date} 查询结果.xlsx"

            output_path = os.path.join(self.output_dir, output_filename)

            wb = Workbook()
            ws = wb.active

            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20

            idx_cell_row = 1
            keys = ["姓名", "部门", "资产名称", "资产编码", "品牌型号", "使用地点", "资产类型"]
            for idx, val in enumerate(keys):
                idx_cell_col = idx + 1
                cell = ws.cell(idx_cell_row, idx_cell_col, value=val)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            for data in data_list:
                assets = data['assets']
                for idx_asset, asset in enumerate(assets):
                    idx_cell_row += 1
                    cell = ws.cell(idx_cell_row, 1, value=data['user'])
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell = ws.cell(idx_cell_row, 2, value=data['department'])
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell = ws.cell(idx_cell_row, 3, value=asset['assetName'])
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell = ws.cell(idx_cell_row, 4, value=asset['assetCoding'])
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell = ws.cell(idx_cell_row, 5, value=f"{asset['supplier']} {asset['specifications']}")
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell = ws.cell(idx_cell_row, 6, value=asset['depositSite'])
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                    ws.row_dimensions[idx_cell_row].height = 30


            # 保存文件
            wb.save(output_path)
            # print(f"✓ 表单已生成: {output_path}")
            return output_path

        except Exception as e:
            self.log.error(f"✗ 填写表单失败: {e}")
            import traceback
            traceback.print_exc()
            return None

def main():
    """ 生成log文件 """

    params = load_param_from_json()
    log = FileLogger("logs", "assetAutomator", level=params.get('log_level', 20))
    log.info("程序启动")

    # 读取需查询数据
    log.info(f"从 {params['in_file_path']} 中解析数据")

    asset_exporter = AssetExporter(
        log.get_logger(),
        params['output_dir']
    )
    data_list = asset_exporter.load_data(params['in_file_path'])
    if type(data_list) == list and len(data_list) > 0:
        log.info(f"成功解析 {len(data_list)} 条记录")
    else:
        log.error("错误：无法解析数据，请确保格式正确（使用制表符分隔）")
        sys.exit(1)

    # 初始化系统
    ams = AssetAutomator(
        params['base_url'],
        params['HEADERS'],
        params['username'],
        params['encrypted_password'],
        log.get_logger()
    )

    # 批量修改资产信息
    for i, data in enumerate(data_list):
        log.info(f"人员{i}. {data['user']} - {data['department']}")
        assets = ams.query_asset_by_department_and_staff(data['department'], data['user'])
        # asset_list.extend(assets)
        data_list[i].update({'assets': assets})

    asset_exporter.write_excel(data_list)
    log.info("程序结束")


# 使用配置文件的方式
def load_param_from_json():
    """从配置文件加载系统配置"""
    config_file = './assetExporter_config.json'
    import json
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except FileNotFoundError:
        print(f"请创建配置文件 {config_file} 后再运行")
        return None



# ===================== 启动程序 =====================
if __name__ == "__main__":
    main()